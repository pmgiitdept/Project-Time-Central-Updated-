#files/views.py
from rest_framework import viewsets, permissions, status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import File, AuditLog, SystemSettings, EmployeeDirectory, DTRFile, DTREntry, Employee, PDFFile
from .serializers import FileSerializer, FileStatusSerializer, AuditLogSerializer, SystemSettingsSerializer, EmployeeDirectorySerializer, DTREntrySerializer, DTRFileSerializer, EmployeeSerializer, PDFFileSerializer
from accounts.permissions import ReadOnlyForViewer, IsOwnerOrAdmin, CanEditStatus, IsAdmin
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.http import FileResponse, Http404, HttpResponse
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from openpyxl import load_workbook, Workbook
import io, re, logging, csv, math, traceback, os, numbers
from accounts.models import User
from reportlab.pdfgen import canvas
from django.db.models import Count, Q
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from .utils import log_action, get_client_ip
from django.core.exceptions import ValidationError
import pandas as pd
from decimal import Decimal, InvalidOperation
from datetime import timedelta
from fuzzywuzzy import fuzz
from django.core.exceptions import FieldDoesNotExist
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def log_action(user, action, status="success", ip_address=None):
    AuditLog.objects.create(
        user=user if user.is_authenticated else None,
        action=action,
        status=status,
        ip_address=ip_address
    )

class FileViewSet(viewsets.ModelViewSet):
    queryset = File.objects.all()
    serializer_class = FileSerializer
    permission_classes = [permissions.IsAuthenticated, CanEditStatus]

    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_parser_classes(self):
        if self.action == "create":
            return [MultiPartParser, FormParser]
        return [JSONParser]

    def get_queryset(self):
        user = self.request.user

        # 🔓 Allow full access for status updates
        if self.action == "update_status" and user.role in ["admin", "viewer"]:
            return File.objects.all().order_by("-uploaded_at")

        if user.role == "client":
            return File.objects.filter(owner=user).order_by("-uploaded_at")

        return File.objects.all().order_by("-uploaded_at")

    def get_permissions(self):
        if self.action == "destroy":
            permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
        elif self.action == "update_status":
            permission_classes = [IsAuthenticated, CanEditStatus]
        else:
            permission_classes = [IsAuthenticated]
        return [p() for p in permission_classes]
    
    def perform_create(self, serializer):
        user = self.request.user
        settings = SystemSettings.objects.first()

        file_obj = serializer.validated_data["file"]

        # 1️⃣ Check file size
        max_size = settings.max_file_size * 1024 * 1024
        if file_obj.size > max_size:
            raise ValidationError(f"File exceeds max size of {settings.max_file_size} MB")

        # 2️⃣ Check allowed types
        ext = file_obj.name.split(".")[-1].lower()
        if ext not in settings.allowed_types:
            raise ValidationError(
                f"File type {ext} not allowed. Allowed: {settings.allowed_types}"
            )

        # 3️⃣ Client overwrite logic
        if user.role == "client":
            existing_file = File.objects.filter(owner=user, file=file_obj.name).first()

            if existing_file:
                # Replace physical file
                existing_file.file.delete(save=False)

                # Save using serializer (update existing instance)
                updated_file = serializer.save(
                    instance=existing_file,
                    owner=user
                )

                log_action(
                    user,
                    f"updated file {updated_file.file.name}",
                    ip_address=get_client_ip(self.request)
                )
                return

        # 4️⃣ Normal upload
        new_file = serializer.save(owner=user)

        log_action(
            user,
            f"uploaded file {new_file.file.name}",
            ip_address=get_client_ip(self.request)
        )

    @action(detail=True, methods=["get"], url_path="download")
    def download(self, request, pk=None):
        file = self.get_object()
        settings = SystemSettings.objects.first()
        
        if settings.require_verification and file.status != "verified":
            return Response({"detail": "File must be verified before download"}, status=403)
        
        if settings.log_downloads:
            log_action(request.user, f"downloaded file {file.file.name}", ip_address=get_client_ip(request))
        
        try:
            return FileResponse(file.file.open(), as_attachment=True, filename=file.file.name)
        except FileNotFoundError:
            raise Http404

    def perform_destroy(self, instance):
        settings = SystemSettings.objects.first()
        
        if settings.auto_archive:
            instance.status = "archived"
            instance.save(update_fields=["status"])
            log_action(self.request.user, f"archived file {instance.file.name}", ip_address=get_client_ip(self.request))
        else:
            file_name = instance.file.name
            super().perform_destroy(instance)
            log_action(self.request.user, f"deleted file {file_name}", ip_address=get_client_ip(self.request))

    @action(detail=True, methods=["get"], url_path="content")
    def get_content(self, request, pk=None):
        file_obj = self.get_object()
        print(f"User: {request.user}, Role: {request.user.role}, File: {file_obj.file.name}")

        if request.user.role not in ["admin", "viewer" , "client"]:
            return Response({"detail": "Forbidden"}, status=403)

        if file_obj.parsed_content:
            return Response({"pages": file_obj.parsed_content})

        file_name = file_obj.file.name.lower()
        try:
            # --- CSV ---
            if file_name.endswith(".csv"):
                file_obj.file.seek(0) 
                file_data = file_obj.file.read()
                decoded_data = file_data.decode("utf-8").splitlines()
                reader = csv.reader(decoded_data)
                return Response({"pages": [{"page_number": 1, "content": list(reader)}]})

            # --- XLSX ---
            elif file_name.endswith(".xlsx"):
                file_obj.file.seek(0)
                file_bytes = io.BytesIO(file_obj.file.read())
                wb = load_workbook(file_bytes, read_only=True)
                ws = wb.active
                rows = [[str(cell) if cell is not None else "" for cell in row] for row in ws.iter_rows(values_only=True)]
                return Response({"pages": [{"page_number": 1, "content": rows}]})

            # --- PDF ---
            elif file_name.endswith(".pdf"):
                import pdfplumber
                pages_data = []

                def is_number(val):
                    try:
                        float(val)
                        return True
                    except ValueError:
                        return False

                file_obj.file.seek(0)
                with pdfplumber.open(file_obj.file) as pdf:
                    for i, page in enumerate(pdf.pages, start=1):
                        page_data = {"page_number": i}

                        text = page.extract_text()
                        if text:
                            page_data["text"] = text

                        lines = text.splitlines() if text else []
                        structured_table = {"main_headers": [], "sub_headers": [], "rows": []}

                        if lines:
                            header_idx = None
                            for idx, line in enumerate(lines):
                                if "Emp." in line and "DUTY" in line:
                                    header_idx = idx
                                    break

                            if header_idx is not None:
                                structured_table["main_headers"] = [
                                    "Emp. No",
                                    "Name",
                                    "Duty (By Days)",
                                    "Late",
                                    "UT",
                                    "Work (By Hrs)",
                                    "Day-Off (By Hours)",
                                    "SH (By Hrs)",
                                    "LH (By Hrs)",
                                    "Day-Off - SH (By Hrs)",
                                    "Day-Off - LH (By Hrs)"
                                ]
                                structured_table["sub_headers"] = [
                                    [""], [""],
                                    ["WRK", "ABS", "LV", "HOL", "RES"],
                                    [""], [""],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                    ["REG", "OT", "ND", "OTND"],
                                ]

                                data_lines = lines[header_idx + 1:]
                                expected_cols = sum(len(group) for group in structured_table["sub_headers"])

                                for dl in data_lines:
                                    parts = dl.split()
                                    if not parts or not parts[0].isdigit():
                                        continue

                                    emp_no = parts[0]
                                    name_parts, numbers = [], []
                                    found_number = False

                                    for p in parts[1:]:
                                        if is_number(p):
                                            found_number = True
                                            numbers.append(f"{float(p):.2f}")
                                        elif not found_number:
                                            name_parts.append(p)

                                    clean_name = " ".join(name_parts).strip()
                                    padded_numbers = numbers + ["0.00"] * (expected_cols - len(numbers))
                                    row = [emp_no, clean_name] + padded_numbers[:expected_cols]
                                    structured_table["rows"].append(row)

                        if structured_table["rows"]:
                            page_data["tables"] = [structured_table]

                        pages_data.append(page_data)

                return Response({"pages": pages_data})

            # --- Images ---
            elif file_name.endswith((".jpg", ".jpeg", ".png")):
                import cv2, numpy as np, easyocr
                file_obj.file.seek(0)
                file_bytes = np.asarray(bytearray(file_obj.file.read()), dtype=np.uint8)
                img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
                if img is None:
                    raise ValueError("Failed to decode image")

                reader = easyocr.Reader(["en"])
                ocr_result = reader.readtext(img)

                rows_dict = {}
                row_threshold = 10
                for (bbox, text, conf) in ocr_result:
                    top = int(bbox[0][1])
                    left = int(bbox[0][0])
                    if not text.strip():
                        continue
                    found = False
                    for key in rows_dict:
                        if abs(key - top) < row_threshold:
                            rows_dict[key].append((left, text))
                            found = True
                            break
                    if not found:
                        rows_dict[top] = [(left, text)]

                table = []
                for top in sorted(rows_dict.keys()):
                    row_words = sorted(rows_dict[top], key=lambda x: x[0])
                    table.append([w[1] for w in row_words])

                return Response({"pages": [{"page_number": 1, "content": table}]})

            else:
                return Response({"detail": "Unsupported file type"}, status=400)

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"detail": f"Failed to read file: {str(e)}"}, status=400)

    @action(detail=True, methods=["patch"], url_path="update-content")
    def update_content(self, request, pk=None):
        file_obj = self.get_object()
        if request.user.role not in ["admin", "viewer"]:
            return Response({"detail": "Forbidden"}, status=403)

        pages = request.data.get("pages")
        content = request.data.get("content") 

        if not pages and not content:
            return Response({"detail": "No content provided"}, status=400)

        if pages and not content:
            content = []
            for page in pages:
                if page.get("tables"):
                    for table in page["tables"]:
                        rows = table.get("rows", [])
                        sub_headers = table.get("sub_headers") or []
                        expected_cols = sum(len(group) if isinstance(group, list) else 1 for group in sub_headers)

                        for row in rows:
                            normalized = []
                            for c in range(expected_cols):
                                if c < len(row):
                                    val = row[c]
                                    try:
                                        normalized.append(float(val))
                                    except (ValueError, TypeError):
                                        normalized.append(str(val or "0.00"))
                                else:
                                    normalized.append("0.00") 
                            content.append(normalized)
                elif page.get("content"): 
                    content.extend(page["content"])

        file_name = file_obj.file.name.lower()

        def format_numeric(val):
            if val is None or val == "":
                return "0.00"
            try:
                num = float(val)
                return f"{num:.2f}"
            except (ValueError, TypeError):
                return str(val)

        try:
            # --- CSV ---
            if file_name.endswith(".csv"):
                from django.core.files.base import ContentFile
                import csv, io
                output = io.StringIO()
                writer = csv.writer(output)
                for row in content:
                    writer.writerow([format_numeric(cell) if isinstance(cell, (int, float, str)) else str(cell) for cell in row])
                file_obj.file.save(file_obj.file.name, ContentFile(output.getvalue()), save=True)

            # --- XLSX ---
            elif file_name.endswith(".xlsx"):
                from openpyxl import Workbook
                from django.core.files.base import ContentFile
                from io import BytesIO

                wb = Workbook()
                ws = wb.active
                for row in content:
                    new_row = []
                    for cell in row:
                        try:
                            new_row.append(float(cell))
                        except (ValueError, TypeError):
                            new_row.append(str(cell or "0.00"))
                    ws.append(new_row)
                stream = BytesIO()
                wb.save(stream)
                file_obj.file.save(file_obj.file.name, ContentFile(stream.getvalue()), save=True)

            # --- PDF ---
            elif file_name.endswith(".pdf"):
                from reportlab.lib.pagesizes import letter
                from reportlab.pdfgen import canvas
                from reportlab.pdfbase.pdfmetrics import stringWidth
                from reportlab.lib.colors import lightgrey, black
                from django.core.files.base import ContentFile
                from io import BytesIO

                buffer = BytesIO()
                p = canvas.Canvas(buffer, pagesize=letter)

                for page in pages:
                    y_offset = 750
                    page_has_content = False

                    if "text" in page and page["text"]:
                        p.setFont("Helvetica", 10)
                        for line in page["text"].splitlines():
                            p.drawString(50, y_offset, line)
                            y_offset -= 15
                            page_has_content = True
                            if y_offset < 50:
                                p.showPage()
                                y_offset = 750
                        p.showPage()

                    if "tables" in page and page["tables"]:
                        for table in page["tables"]:
                            main_headers = table.get("main_headers", [])
                            sub_headers = table.get("sub_headers", [])
                            rows = table.get("rows", [])

                            flat_sub_headers = []
                            for group in sub_headers:
                                flat_sub_headers.extend(group if isinstance(group, list) else [group])

                            all_rows = [main_headers] + [flat_sub_headers] + rows
                            if not all_rows:
                                continue

                            num_cols = max(len(r) for r in all_rows)
                            col_widths = []
                            row_height = 20

                            for col_idx in range(num_cols):
                                max_width = max(
                                    stringWidth(format_numeric(row[col_idx]) if col_idx < len(row) else "0.00", "Helvetica", 8)
                                    for row in all_rows
                                )
                                col_widths.append(max_width + 20)

                            x_offset, y_start = 50, y_offset

                            # Main headers
                            x = x_offset
                            for j, h in enumerate(main_headers):
                                span = len(sub_headers[j]) if j < len(sub_headers) else 1
                                span_width = sum(col_widths[j:j+span])
                                p.setFillColor(lightgrey)
                                p.rect(x, y_start, span_width, -row_height, fill=1, stroke=1)
                                p.setFillColor(black)
                                p.setFont("Helvetica-Bold", 8)
                                p.drawCentredString(x + span_width / 2, y_start - row_height + 15, h)
                                x += span_width
                            y_start -= row_height

                            # Sub headers
                            x = x_offset
                            for group in sub_headers:
                                for sh in (group if isinstance(group, list) else [group]):
                                    width = col_widths[sub_headers.index(group)]
                                    p.rect(x, y_start, width, -row_height, fill=0, stroke=1)
                                    p.setFont("Helvetica-Bold", 8)
                                    p.drawCentredString(x + width / 2, y_start - row_height + 15, sh)
                                    x += width
                            y_start -= row_height

                            # Table rows
                            for row in rows:
                                if y_start < 50:
                                    p.showPage()
                                    y_start = 750
                                x = x_offset
                                for j in range(num_cols):
                                    text = format_numeric(row[j]) if j < len(row) else "0.00"
                                    p.rect(x, y_start, col_widths[j], -row_height, fill=0, stroke=1)
                                    p.setFont("Helvetica", 8)
                                    p.drawString(x + 2, y_start - row_height + 15, text)
                                    x += col_widths[j]
                                y_start -= row_height

                            y_offset = y_start
                        p.showPage()

                    if not page_has_content:
                        p.setFont("Helvetica", 10)
                        p.drawString(50, 750, "No content available")
                        p.showPage()

                p.save()
                buffer.seek(0)
                file_obj.file.save(file_obj.file.name, ContentFile(buffer.read()), save=True)

            # --- Images ---
            elif file_name.endswith((".jpg", ".jpeg", ".png")):
                import cv2
                import numpy as np
                from django.core.files.base import ContentFile

                rows = content
                cell_width, cell_height = 200, 50
                font, font_scale, thickness = cv2.FONT_HERSHEY_SIMPLEX, 0.7, 1

                n_rows = len(rows)
                n_cols = max(len(r) for r in rows) if rows else 0
                if n_rows == 0 or n_cols == 0:
                    return Response({"detail": "No content to update"}, status=400)

                img_height = n_rows * cell_height + 2
                img_width = n_cols * cell_width + 2
                img = np.ones((img_height, img_width, 3), dtype=np.uint8) * 255

                for i, row in enumerate(rows):
                    for j, cell in enumerate(row):
                        x1, y1 = j * cell_width, i * cell_height
                        x2, y2 = x1 + cell_width, y1 + cell_height
                        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 0), 1)
                        text = str(cell)
                        (tw, th), _ = cv2.getTextSize(text, font, font_scale, thickness)
                        text_x = x1 + (cell_width - tw) // 2
                        text_y = y1 + (cell_height + th) // 2
                        cv2.putText(img, text, (text_x, text_y), font, font_scale, (0, 0, 0), thickness)

                _, buffer = cv2.imencode(".png", img)
                file_obj.file.save(file_obj.file.name, ContentFile(buffer.tobytes()), save=True)

            else:
                return Response({"detail": "Unsupported file type"}, status=400)

            file_obj.parsed_content = pages or content
            file_obj.save(update_fields=["parsed_content"])
            return Response({"detail": "Content updated successfully"})

        except Exception as e:
            return Response({"detail": f"Failed to save content: {str(e)}"}, status=400)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    files_pending = DTRFile.objects.filter(status="pending").count()
    files_rejected = DTRFile.objects.filter(status="rejected").count()
    files_approved = DTRFile.objects.filter(status="verified").count()
    active_users = User.objects.filter(is_active=True).count()
    
    return Response({
        "filesPending": files_pending,
        "filesApproved": files_approved,
        "filesRejected": files_rejected,
        "activeUsers": active_users
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_files_report(request):
    format = request.GET.get("format", "csv")
    files = File.objects.all().order_by("-uploaded_at")
    
    if format == "csv":
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = 'attachment; filename="files_report.csv"'
        writer = csv.writer(response)
        writer.writerow(["ID", "Filename", "Owner", "Status", "Uploaded At"])
        for f in files:
            writer.writerow([f.id, f.file.name, f.owner.username, f.status, f.uploaded_at])
        return response
    
    elif format == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="files_report.pdf"'
        p = canvas.Canvas(response)
        y = 800
        p.drawString(50, y, "Files Report")
        y -= 25
        for f in files:
            p.drawString(50, y, f"{f.id} | {f.file.name} | {f.owner.username} | {f.status} | {f.uploaded_at}")
            y -= 20
        p.showPage()
        p.save()
        return response
    
    else:
        return Response({"detail": "Unsupported format"}, status=400)
    
class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all().order_by("-timestamp")
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdmin]  

class SystemSettingsViewSet(viewsets.ModelViewSet):
    queryset = SystemSettings.objects.all()
    serializer_class = SystemSettingsSerializer
    permission_classes = [IsAdmin] 

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def file_stats(request):
    period = request.query_params.get("period", "day")  

    if period == "month":
        trunc = TruncMonth("uploaded_at")
    elif period == "week":
        trunc = TruncWeek("uploaded_at")
    else:
        trunc = TruncDay("uploaded_at")

    stats = (
        File.objects
        .annotate(period=trunc)
        .values("period")
        .annotate(
            pending=Count("id", filter=Q(status="pending")),
            verified=Count("id", filter=Q(status="verified")),
            rejected=Count("id", filter=Q(status="rejected")),
        )
        .order_by("period")
    )

    return Response(stats)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def rejected_files(request):
    user = request.user
    files = File.objects.filter(owner=user, status="rejected").order_by("-uploaded_at")
    serializer = FileSerializer(files, many=True)
    return Response(serializer.data)

def clean_value(value):
    """Convert NaN or missing values to None for JSON storage."""
    if pd.isna(value) or (isinstance(value, float) and math.isnan(value)):
        return None
    return value

@api_view(['POST'])
@permission_classes([IsAdminUser])
def upload_employee_excel(request):
    file = request.FILES.get('file')
    if not file:
        return Response({"detail": "No file uploaded."}, status=400)
    
    try:
        df = pd.read_excel(file, header=0, dtype=str)
        df.columns = [col.strip() for col in df.columns]  
        df_cleaned = df.applymap(clean_value)

        added_count = 0
        updated_count = 0

        def to_float(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        for _, row in df_cleaned.iterrows():
            employee_code_raw = row.get('Employee Code')
            employee_code = str(employee_code_raw).strip().zfill(5) if employee_code_raw else ""

            data = {
                'employee_code': employee_code,
                'employee_name': str(row.get('EmployeeName', '')).strip(),
                'total_hours': to_float(row.get('Total Hours')),
                'nd_reg_hrs': to_float(row.get('ND Reg Hrs')),
                'absences': to_float(row.get('Absences')),
                'tardiness': to_float(row.get('Tardiness')),
                'undertime': to_float(row.get('Undertime')),
                'ot_regular': to_float(row.get('OTRegular')),
                'nd_ot_reg': to_float(row.get('ND OT Reg')),
                'ot_restday': to_float(row.get('OT Restday')),
                'nd_restday': to_float(row.get('ND Restday')),
                'ot_rest_excess': to_float(row.get('OT RestExcess')),
                'nd_rest_excess': to_float(row.get('ND Restday Excess')),
                'ot_special_hday': to_float(row.get('OTSpecialHday')),
                'nd_special_hday': to_float(row.get('ND SpecialHday')),
                'ot_shday_excess': to_float(row.get('OT SHdayExcess')),
                'nd_shday_excess': to_float(row.get('ND SHday Excess')),
                'ot_legal_holiday': to_float(row.get('OT LegalHoliday')),
                'special_holiday': to_float(row.get('Special Holiday')),
                'ot_leghol_excess': to_float(row.get('OTLegHol Excess')),
                'nd_leghol_excess': to_float(row.get('ND LegHol Excess')),
                'ot_sh_on_rest': to_float(row.get('OT SHday on Rest')),
                'nd_sh_on_rest': to_float(row.get('ND SH on Rest')),
                'ot_sh_on_rest_excess': to_float(row.get('OT SH on Rest Excess')),
                'nd_sh_on_rest_excess': to_float(row.get('ND SH on Rest Excess')),
                'leg_h_on_rest_day': to_float(row.get('LegH on Rest Day')),
                'nd_leg_h_on_restday': to_float(row.get('ND LegH on Restday')),
                'ot_leg_h_on_rest_excess': to_float(row.get('OT LegH on Rest Excess')),
                'nd_leg_h_on_rest_excess': to_float(row.get('ND LegH on Rest Excess')),
                'vacleave_applied': to_float(row.get('VacLeave_Applied')),
                'sickleave_applied': to_float(row.get('SickLeave_Applied')),
                'back_pay_vl': to_float(row.get('Back Pay VL')),
                'back_pay_sl': to_float(row.get('Back Pay SL')),
                'ot_regular_excess': to_float(row.get('OTRegular Excess')),
                'nd_ot_reg_excess': to_float(row.get('ND OT Reg Excess')),
                'legal_holiday': to_float(row.get('Legal Holiday')),
                'nd_legal_holiday': to_float(row.get('ND Legal Holiday')),
                'overnight_rate': to_float(row.get('Overnight Rate')),
                'project': str(row.get('PROJECT', '')).strip(),
            }

            if employee_code:
                obj, created = EmployeeDirectory.objects.update_or_create(
                    employee_code=employee_code,
                    defaults=data
                )
                if created:
                    added_count += 1
                else:
                    updated_count += 1
            else:
                obj = EmployeeDirectory.objects.filter(employee_name=data['employee_name'], employee_code="").first()
                if obj:
                    for key, value in data.items():
                        setattr(obj, key, value)
                    obj.save()
                    updated_count += 1
                else:
                    EmployeeDirectory.objects.create(**data)
                    added_count += 1

        return Response({
            "detail": f"{added_count} new employees added, {updated_count} employees updated."
        })

    except Exception as e:
        return Response({"detail": str(e)}, status=400)
    
@api_view(['GET'])
@permission_classes([IsAdminUser])
def list_employees(request):
    employees = EmployeeDirectory.objects.all().order_by("id")
    serializer = EmployeeDirectorySerializer(employees, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAdminUser])
def add_employee(request):
    employee_code = request.data.get("employee_code")
    employee_name = request.data.get("employee_name")

    if not employee_code or not employee_name:
        return Response({"detail": "Employee code and name are required."}, status=400)

    employee_code = str(employee_code).strip().zfill(5)
    employee_name = employee_name.strip()

    obj, created = EmployeeDirectory.objects.get_or_create(
        employee_code=employee_code,
        defaults={"employee_name": employee_name}
    )

    if not created:
        return Response({"detail": "Employee already exists."}, status=400)

    return Response({"detail": "Employee added successfully!"})

@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def delete_employee(request, employee_code):
    employee_code_str = str(employee_code)
    padded_code = employee_code_str.zfill(5)

    # Fetch possible matches safely
    matches = EmployeeDirectory.objects.filter(
        Q(employee_code=padded_code) | Q(employee_code=employee_code_str)
    )

    count = matches.count()
    if count == 0:
        return Response(
            {"detail": "Employee not found."},
            status=status.HTTP_404_NOT_FOUND
        )
    elif count > 1:
        return Response(
            {"detail": f"Duplicate entries found for employee code {employee_code_str}. Please resolve duplicates first."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Proceed to delete the single match
    employee = matches.first()
    employee.delete()

    return Response(
        {"detail": "Employee deleted successfully."},
        status=status.HTTP_200_OK
    )


@api_view(['PUT'])
@permission_classes([IsAdminUser])
def update_employee(request, employee_code):
    employee_code_str = str(employee_code)
    padded_code = employee_code_str.zfill(5)
    
    try:
        employee = EmployeeDirectory.objects.get(
            Q(employee_code=padded_code) | Q(employee_code=employee_code_str)
        )
    except EmployeeDirectory.DoesNotExist:
        return Response({"detail": "Employee not found"}, status=404)

    numeric_fields = [
        "total_hours", "nd_reg_hrs", "absences", "tardiness", "undertime",
        "ot_regular", "nd_ot_reg", "ot_restday", "nd_restday", "ot_rest_excess",
        "nd_rest_excess", "ot_special_hday", "nd_special_hday", "ot_shday_excess",
        "nd_shday_excess", "ot_legal_holiday", "special_holiday", "ot_leghol_excess",
        "nd_leghol_excess", "ot_sh_on_rest", "nd_sh_on_rest", "ot_sh_on_rest_excess",
        "nd_sh_on_rest_excess", "leg_h_on_rest_day", "nd_leg_h_on_restday",
        "ot_leg_h_on_rest_excess", "nd_leg_h_on_rest_excess", "vacleave_applied",
        "sickleave_applied", "back_pay_vl", "back_pay_sl", "ot_regular_excess",
        "nd_ot_reg_excess", "legal_holiday", "nd_legal_holiday", "overnight_rate",
    ]

    for field, value in request.data.items():
        if hasattr(employee, field):
            if field in numeric_fields:
                if value in ["", None]:
                    setattr(employee, field, None)
                else:
                    try:
                        setattr(employee, field, Decimal(value))
                    except InvalidOperation:
                        setattr(employee, field, None)
            else:
                setattr(employee, field, value)

    employee.save()
    return Response({"detail": "Employee updated successfully"})

def safe_number(val, default=0):
    """Convert pandas/Excel values into a safe float or default."""
    if pd.isna(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_string(val):
    """Convert to string if not NaN, else return None."""
    if pd.isna(val):
        return None
    return str(val).strip()

class DTRFileViewSet(viewsets.ModelViewSet):
    queryset = DTRFile.objects.all().order_by("-uploaded_at")
    serializer_class = DTRFileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        if hasattr(user, "role") and user.role in ["admin", "viewer"]:
            return DTRFile.objects.all().order_by("-uploaded_at")

        return DTRFile.objects.filter(uploaded_by=user).order_by("-uploaded_at")

    def perform_create(self, serializer):
        # 1️⃣ Save the file and capture the instance
        dtr_file = serializer.save(uploaded_by=self.request.user)

        # 2️⃣ Log upload action
        log_action(
            user=self.request.user,
            action=f"Uploaded DTR file: {dtr_file.file.name}",
            ip_address=self.request.META.get("REMOTE_ADDR")
        )

    @action(detail=True, methods=["post"])
    def parse(self, request, pk=None):
        dtr_file = self.get_object()
        file_path = dtr_file.file.path
        dtr_file.entries.all().delete()  # clear existing entries

        def parse_excel_date(val):
            """Safely parse a date from Excel cell (merged cells, strings, numbers)"""
            if pd.isna(val):
                return None
            # Excel serial number
            if isinstance(val, (int, float)):
                try:
                    return (pd.to_datetime("1899-12-30") + pd.to_timedelta(val, "D")).date()
                except:
                    return None
            # Already datetime
            if isinstance(val, (pd.Timestamp, datetime)):
                return val.date()
            # String parsing
            try:
                s = str(val).strip()
                # Remove prefix like "Start Date: "
                if ":" in s:
                    s = s.split(":")[-1]
                return pd.to_datetime(s, errors="coerce").date()
            except:
                return None

        try:
            df = pd.read_excel(file_path, header=None)

            # Extract start_date from D9:E9 (row 8, cols 3-4)
            start_date_val = None
            for val in df.iloc[8, 3:5]:
                start_date_val = parse_excel_date(val)
                if start_date_val:
                    break

            # Extract end_date from D10:E10 (row 9, cols 3-4)
            end_date_val = None
            for val in df.iloc[9, 3:5]:
                end_date_val = parse_excel_date(val)
                if end_date_val:
                    break

            # Save to model
            dtr_file.start_date = start_date_val
            dtr_file.end_date = end_date_val
            dtr_file.save()

            # Employee rows start from row 15 (index 14)
            employee_df = df.iloc[14:, :]

            for _, row in employee_df.iterrows():
                name = row[2] if len(row) > 2 else None
                emp_no = row[3] if len(row) > 3 else None
                if pd.isna(name) or pd.isna(emp_no):
                    continue

                emp_code = str(emp_no).strip()
                if emp_code.startswith("PM"):
                    emp_code = emp_code[2:]
                emp_code = "".join(filter(str.isdigit, emp_code))

                daily_data = {}
                if dtr_file.start_date:
                    for idx, col in enumerate(range(7, 23)):  # H → W
                        day = dtr_file.start_date + timedelta(days=idx)
                        val = row[col] if col < len(row) else None
                        daily_data[str(day)] = None if pd.isna(val) else val

                DTREntry.objects.create(
                    dtr_file=dtr_file,
                    full_name=safe_string(name),
                    employee_no=emp_code,
                    position=safe_string(row[4]) if len(row) > 4 else None,
                    shift=safe_string(row[5]) if len(row) > 5 else None,
                    time=safe_string(row[6]) if len(row) > 6 else None,
                    daily_data=daily_data,
                    total_days=safe_number(row[23]) if len(row) > 23 else 0,
                    total_hours=safe_number(row[24]) if len(row) > 24 else 0,
                    undertime_minutes=safe_number(row[25]) if len(row) > 25 else 0,
                    regular_ot=safe_number(row[26]) if len(row) > 26 else 0,
                    legal_holiday=safe_number(row[27]) if len(row) > 27 else 0,
                    unworked_reg_holiday=safe_number(row[28]) if len(row) > 28 else 0,
                    special_holiday=safe_number(row[29]) if len(row) > 29 else 0,
                    night_diff=safe_number(row[30]) if len(row) > 30 else 0,
                )

            return Response({"message": "DTR file parsed successfully."})

        except Exception as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"])
    def content(self, request, pk=None):
        dtr_file = self.get_object()
        entries = dtr_file.entries.all()
        serializer = DTREntrySerializer(entries, many=True)
        return Response({
            "start_date": dtr_file.start_date,
            "end_date": dtr_file.end_date,
            "rows": serializer.data
        })
    
    @action(
        detail=True,
        methods=["get", "patch"],
        url_path="status",
        parser_classes=[JSONParser],
        permission_classes=[IsAuthenticated, CanEditStatus],
    )
    def status(self, request, pk=None):
        file = self.get_object()

        # ---------- GET ----------
        if request.method == "GET":
            return Response({
                "id": file.id,
                "status": file.status,
                "rejection_reason": file.rejection_reason
            })

        # ---------- PATCH ----------
        previous_status = file.status

        serializer = FileStatusSerializer(file, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        new_status = serializer.validated_data.get("status")
        rejection_reason = serializer.validated_data.get("rejection_reason")

        file_name = os.path.basename(file.file.name) if file.file and file.file.name else f"Manual DTR ({file.id})"

        log_action(
            user=request.user,
            action=(
                f"Updated status of file '{file_name}' "
                f"from '{previous_status}' to '{new_status}'"
                + (
                    f" | Rejection reason: {rejection_reason}"
                    if new_status == "rejected" and rejection_reason
                    else ""
                )
            ),
            status="success",
            ip_address=get_client_ip(request)
        )

        return Response(serializer.data)
    
    @action(detail=True, methods=["get"], url_path="download")
    def download(self, request, pk=None):
        file = self.get_object()
        settings = SystemSettings.objects.first()
        
        if settings.require_verification and file.status != "verified":
            return Response({"detail": "File must be verified before download"}, status=403)
        
        if settings.log_downloads:
            log_action(request.user, f"downloaded file {file.file.name}", ip_address=get_client_ip(request))
        
        try:
            return FileResponse(file.file.open(), as_attachment=True, filename=file.file.name)
        except FileNotFoundError:
            raise Http404

    @action(detail=False, methods=["post"], url_path="sync-all")
    def sync_all_files(self, request):
        """
        Sync only verified DTR files to EmployeeDirectory.
        Optional request data:
            start_date, end_date -> filter which DTR files to consider
        """
        start_date_str = request.data.get("start_date")
        end_date_str = request.data.get("end_date")

        start_date = pd.to_datetime(start_date_str).date() if start_date_str else None
        end_date = pd.to_datetime(end_date_str).date() if end_date_str else None

        dtr_files = DTRFile.objects.filter(status="verified").order_by("start_date")

        if start_date:
            dtr_files = dtr_files.filter(end_date__gte=start_date)
        if end_date:
            dtr_files = dtr_files.filter(start_date__lte=end_date)

        if not dtr_files.exists():
            return Response(
                {"detail": "No verified DTR files found for syncing."},
                status=status.HTTP_400_BAD_REQUEST
            )

        aggregated = {}
        for dtr_file in dtr_files:
            uploader = dtr_file.uploaded_by 
            project_name = (
                f"{uploader.first_name} {uploader.last_name}".strip()
                if uploader and (uploader.first_name or uploader.last_name)
                else uploader.username if uploader else "Unknown Project"
            )

            for entry in dtr_file.entries.all():
                code = str(entry.employee_no).zfill(5) if entry.employee_no else None
                name = entry.full_name.strip() if entry.full_name else None
                if not code or not name:
                    continue

                if code not in aggregated:
                    aggregated[code] = {
                        "employee_name": name,
                        "project": project_name,
                        "total_hours": 0,
                        "undertime": 0,
                        "ot_regular": 0,
                        "legal_holiday": 0,
                        "special_holiday": 0,
                        "nd_reg_hrs": 0,
                        "date_covered_start": dtr_file.start_date,
                        "date_covered_end": dtr_file.end_date,
                    }

                aggregated[code]["total_hours"] += entry.total_hours or 0
                aggregated[code]["undertime"] += entry.undertime_minutes or 0
                aggregated[code]["ot_regular"] += entry.regular_ot or 0
                aggregated[code]["legal_holiday"] += entry.legal_holiday or 0
                aggregated[code]["special_holiday"] += entry.special_holiday or 0
                aggregated[code]["nd_reg_hrs"] += entry.night_diff or 0

                if dtr_file.start_date and (
                    aggregated[code]["date_covered_start"] is None
                    or dtr_file.start_date < aggregated[code]["date_covered_start"]
                ):
                    aggregated[code]["date_covered_start"] = dtr_file.start_date

                if dtr_file.end_date and (
                    aggregated[code]["date_covered_end"] is None
                    or dtr_file.end_date > aggregated[code]["date_covered_end"]
                ):
                    aggregated[code]["date_covered_end"] = dtr_file.end_date

        created, updated = 0, 0
        for code, data in aggregated.items():
            date_covered = None
            if data.get("date_covered_start") and data.get("date_covered_end"):
                start_str = data["date_covered_start"].strftime("%b %d, %Y")
                end_str = data["date_covered_end"].strftime("%b %d, %Y")
                date_covered = f"{start_str} → {end_str}"

            # 🔧 Handle duplicates safely
            matches = EmployeeDirectory.objects.filter(employee_code=code)
            if matches.count() > 1:
                keep = matches.first()
                matches.exclude(id=keep.id).delete()
                emp = keep
                is_created = False
            else:
                emp, is_created = EmployeeDirectory.objects.update_or_create(
                    employee_code=code,
                    defaults={
                        **{k: v for k, v in data.items() if k not in ["date_covered_start", "date_covered_end"]},
                        "date_covered": date_covered,
                    },
                )

            if is_created:
                created += 1
            else:
                updated += 1

        return Response({
            "detail": f"Synced {created} new, {updated} updated (verified only)."
        })

        
    @action(detail=True, methods=["post"], url_path="update-rows")
    def update_rows(self, request, pk=None):
        dtr_file = self.get_object()
        rows = request.data.get("rows", [])

        for row in rows:
            entry_id = row.get("id")
            if entry_id:
                try:
                    entry = dtr_file.entries.get(id=entry_id)
                except DTREntry.DoesNotExist:
                    continue
            else:
                entry = DTREntry(dtr_file=dtr_file)

            entry.full_name = row.get("full_name", entry.full_name)
            entry.employee_no = row.get("employee_no", entry.employee_no)

            entry.daily_data = row.get("daily_data", entry.daily_data)

            entry.total_days = row.get("total_days", entry.total_days)
            entry.total_hours = row.get("total_hours", entry.total_hours)
            entry.regular_ot = row.get("regular_ot", entry.regular_ot)
            entry.legal_holiday = row.get("legal_holiday", entry.legal_holiday)
            entry.unworked_reg_holiday = row.get("unworked_reg_holiday", entry.unworked_reg_holiday)
            entry.special_holiday = row.get("special_holiday", entry.special_holiday)
            entry.night_diff = row.get("night_diff", entry.night_diff)
            entry.undertime_minutes = row.get("undertime_minutes", entry.undertime_minutes)

            entry.save()

        return Response({"message": "DTR entries updated successfully!"}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=["post"], url_path="manual")
    def manual(self, request):
        data = request.data

        start_date = pd.to_datetime(data.get("start_date")).date()
        end_date = pd.to_datetime(data.get("end_date")).date()
        rows = data.get("rows", [])

        if not rows:
            return Response({"detail": "No rows provided"}, status=400)

        # 1️⃣ Create DTRFile (same model Excel uses)
        dtr_file = DTRFile.objects.create(
            uploaded_by=request.user,
            start_date=start_date,
            end_date=end_date,
            status="pending",  # same lifecycle
        )

        # 🔹 Log manual creation for audit
        log_action(
            user=request.user,
            action=f"Created manual DTR: {dtr_file.id} ({start_date} to {end_date})",
            ip_address=request.META.get("REMOTE_ADDR")
        )

        # 2️⃣ Create DTREntry rows (same fields parse() fills)
        for row in rows:
            DTREntry.objects.create(
                dtr_file=dtr_file,
                full_name=row.get("full_name"),
                employee_no=row.get("employee_no"),
                position=row.get("position"),
                shift=row.get("shift"),
                time=row.get("time"),
                daily_data=row.get("daily_data", {}),
                total_days=row.get("total_days"),
                total_hours=row.get("total_hours"),
                undertime_minutes=row.get("undertime_minutes"),
                regular_ot=row.get("regular_ot"),
                legal_holiday=row.get("legal_holiday"),
                unworked_reg_holiday=row.get("unworked_reg_holiday"),
                special_holiday=row.get("special_holiday"),
                night_diff=row.get("night_diff"),
            )

        return Response(
            {"message": "Manual DTR created successfully", "id": dtr_file.id},
            status=201
        )

    @action(detail=True, methods=["get"], url_path="export")
    def export(self, request, pk=None):
        dtr = self.get_object()

        # Return original uploaded file if exists
        if dtr.file:
            response = HttpResponse(
                dtr.file.open("rb"),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{os.path.basename(dtr.file.name)}"'
            return response

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"

        # ----------------------------
        # 🎨 Styles
        # ----------------------------
        bold = Font(name="Leelawadee", bold=True)
        title_font = Font(name="Leelawadee", size=14, bold=True)
        subtitle_font = Font(name="Leelawadee", size=12, bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        # Daily value fills
        blank_fill = PatternFill("solid", fgColor="4F81BD")   # Blue
        absent_fill = PatternFill("solid", fgColor="FF0000") # Red
        dayoff_fill = PatternFill("solid", fgColor="00B050") # Green

        # Daily data header color
        daily_header_fill = PatternFill("solid", fgColor="4F81BD")
        daily_header_font = Font(name="Leelawadee", color="FFFFFF", bold=True)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        total_columns = 6 + (dtr.end_date - dtr.start_date).days + 1 + 7

        # ----------------------------
        # 🧱 HEADER (Logo + Main/Sub Headers)
        # ----------------------------
        table_first_col = 2      # Table starts at column A
        logo_cols_span = 2       # Logo width in columns
        logo_rows_span = 5       # Logo height in rows
        header_start_col = table_first_col + logo_cols_span
        header_end_col = total_columns  # Total table columns

        from openpyxl.drawing.image import Image as XLImage
        logo_path = os.path.join("media", "pmgi.png")

        # ----------------------------
        # Merge logo area
        # ----------------------------
        ws.merge_cells(
            start_row=1, start_column=table_first_col,
            end_row=logo_rows_span, end_column=table_first_col + logo_cols_span - 1
        )

        # ----------------------------
        # Insert logo and scale to exact merged cell area
        # ----------------------------
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            
            # Calculate merged cell width in pixels
            cell_width = sum(
                (ws.column_dimensions[get_column_letter(c)].width or 8) for c in range(table_first_col, table_first_col + logo_cols_span)
            ) * 7.7  # approximate pixels

            # Calculate merged cell height in pixels
            cell_height = sum(
                (ws.row_dimensions[r].height or 18) for r in range(1, logo_rows_span + 1)
            ) * 1.09  # approximate pixels

            # Slightly scale up to fill merged area
            scale_factor = 1.06
            img.width = int(cell_width * scale_factor)
            img.height = int(cell_height * scale_factor)

            # Insert image at top-left of merged cells
            ws.add_image(img, ws.cell(row=1, column=table_first_col).coordinate)

        # ----------------------------
        # Main header (rows 1-2)
        # ----------------------------
        ws.merge_cells(
            start_row=1, start_column=header_start_col,
            end_row=2, end_column=header_end_col
        )
        main_header_cell = ws.cell(row=1, column=header_start_col, value="PROFESSIONAL MAINTENANCE GROUP, INC.")
        main_header_cell.font = title_font
        main_header_cell.alignment = center

        # ----------------------------
        # Subheader (row 3)
        # ----------------------------
        ws.merge_cells(
            start_row=3, start_column=header_start_col,
            end_row=4, end_column=header_end_col
        )
        subheader_cell = ws.cell(row=3, column=header_start_col, value="ATTENDANCE SUMMARY FORM")
        subheader_cell.font = subtitle_font
        subheader_cell.alignment = center

        # ----------------------------
        # OPS-FRM code (row 4)
        # ----------------------------
        ws.merge_cells(
            start_row=5, start_column=header_start_col,
            end_row=5, end_column=header_end_col - 2
        )
        ops_cell = ws.cell(row=5, column=header_start_col, value="OPS-FRM009-01-100125")
        ops_cell.alignment = center

        # ----------------------------
        # Edition (row 4, rightmost columns)
        # ----------------------------
        col_edition = header_end_col - 1
        col_number = header_end_col
        ws.cell(row=5, column=col_edition, value="Edition").font = bold
        ws.cell(row=5, column=col_number, value="3").font = bold
        ws.cell(row=5, column=col_edition).alignment = center
        ws.cell(row=5, column=col_number).alignment = center

        # ----------------------------
        # Apply borders to all header cells (including logo area)
        # ----------------------------
        for row in ws.iter_rows(min_row=1, max_row=logo_rows_span,
                                min_col=table_first_col, max_col=header_end_col):
            for cell in row:
                cell.border = border

        # ----------------------------
        # 🧾 META INFO (Updated)
        # ----------------------------
        meta_row = 7
        uploader = dtr.uploaded_by.username if dtr.uploaded_by else "N/A"

        # PROJECT label (column D)
        ws[f"D{meta_row}"] = "PROJECT:"
        ws[f"D{meta_row}"].font = bold
        ws[f"D{meta_row}"].alignment = right  # label stays left-aligned
        ws[f"D{meta_row}"].border = border

        # Project name / uploaded by (merge E and F)
        start_col = 5  # column E
        end_col = 6    # merge E and F
        ws.merge_cells(start_row=meta_row, start_column=start_col, end_row=meta_row, end_column=end_col)
        ws[f"E{meta_row}"] = uploader
        ws[f"E{meta_row}"].font = bold
        ws[f"E{meta_row}"].alignment = Alignment(horizontal="center", vertical="center")  # center in merged cell

        # Apply border to all merged cells
        for col in range(start_col, end_col + 1):
            ws.cell(row=meta_row, column=col).border = border

        # FROM label and date (next row)
        from_text = dtr.start_date.strftime("%B %d, %Y")  # "September 16, 2025"
        ws[f"D{meta_row+1}"] = "FROM:"
        ws[f"D{meta_row+1}"].font = bold
        ws[f"D{meta_row+1}"].alignment = right
        ws[f"D{meta_row+1}"].border = border

        # Merge E and F for the date
        start_col = 5
        end_col = 6
        ws.merge_cells(start_row=meta_row+1, start_column=start_col, end_row=meta_row+1, end_column=end_col)
        ws[f"E{meta_row+1}"] = from_text
        ws[f"E{meta_row+1}"].font = bold
        ws[f"E{meta_row+1}"].alignment = Alignment(horizontal="center", vertical="center")  # center text in merged cell

        # Apply border to all merged cells
        for col in range(start_col, end_col + 1):
            ws.cell(row=meta_row+1, column=col).border = border

        # TO label and date (row after FROM)
        to_text = dtr.end_date.strftime("%B %d, %Y")  # "September 20, 2025"
        ws[f"D{meta_row+2}"] = "TO:"
        ws[f"D{meta_row+2}"].font = bold
        ws[f"D{meta_row+2}"].alignment = right
        ws[f"D{meta_row+2}"].border = border

        # Merge E and F for TO date
        ws.merge_cells(start_row=meta_row+2, start_column=start_col, end_row=meta_row+2, end_column=end_col)
        ws[f"E{meta_row+2}"] = to_text
        ws[f"E{meta_row+2}"].font = bold
        ws[f"E{meta_row+2}"].alignment = Alignment(horizontal="center", vertical="center")  # center text

        # Apply border to all merged cells
        for col in range(start_col, end_col + 1):
            ws.cell(row=meta_row+2, column=col).border = border

        # ----------------------------
        # 📅 DATES
        # ----------------------------
        dates = []
        cur = dtr.start_date
        while cur <= dtr.end_date:
            dates.append(cur)
            cur += timedelta(days=1)

        table_start = meta_row + 6
        headers_left = ["NO.", "EMPLOYEE NAME", "PM NO.", "POSITION", "SHIFT", "TIME"]
        headers_right = ["Total Days", "Total Hours", "Undertime", "OT", "Legal Hol", "Special Hol", "Night Diff"]

        # Left headers
        for i, header in enumerate(headers_left, start=1):
            ws.merge_cells(start_row=table_start, start_column=i, end_row=table_start+1, end_column=i)
            cell = ws.cell(row=table_start, column=i, value=header)
            cell.font = bold
            cell.alignment = center
            cell.border = border

        # Right headers
        col_idx = 6 + len(dates) + 1
        for header in headers_right:
            ws.merge_cells(start_row=table_start, start_column=col_idx, end_row=table_start+1, end_column=col_idx)
            cell = ws.cell(row=table_start, column=col_idx, value=header)
            cell.font = bold
            cell.alignment = center
            cell.border = border
            col_idx += 1

        # ----------------------------
        # Daily headers
        # ----------------------------
        daily_start_col = 7
        for d in dates:
            fill = daily_header_fill
            font = daily_header_font
            if d.weekday() == 6:  # Sunday
                fill = PatternFill("solid", fgColor="FFFF00")
                font = Font(name="Leelawadee", color="000000", bold=True)

            # Day name
            cell = ws.cell(row=table_start, column=daily_start_col, value=d.strftime("%a"))
            cell.font = font
            cell.fill = fill
            cell.alignment = center
            cell.border = border

            # Day number
            cell = ws.cell(row=table_start+1, column=daily_start_col, value=d.day)
            cell.font = font
            cell.fill = fill
            cell.alignment = center
            cell.border = border

            daily_start_col += 1

        # ----------------------------
        # Helper function
        # ----------------------------
        def is_number(val):
            try:
                float(val)
                return True
            except (TypeError, ValueError):
                return False

        # ----------------------------
        # Data rows
        # ----------------------------
        row_idx = table_start + 2
        daily_col_start = 7
        daily_col_end = daily_col_start + len(dates) - 1

        for idx, entry in enumerate(dtr.entries.all(), start=1):
            daily_vals = [entry.daily_data.get(d.strftime("%Y-%m-%d"), "") for d in dates]

            # Compute numeric-only totals
            numeric_vals = [float(v) for v in daily_vals if is_number(v)]
            total_days = len(numeric_vals)
            total_hours = sum(numeric_vals)
            regular_ot = sum(max(v - 8, 0) for v in numeric_vals)

            # Values to write
            values = (
                [idx, entry.full_name, entry.employee_no, entry.position, entry.shift, entry.time]
                + daily_vals
                + [
                    total_days,
                    total_hours,
                    entry.undertime_minutes or 0,
                    regular_ot,
                    entry.legal_holiday or 0,
                    entry.special_holiday or 0,
                    entry.night_diff or 0,
                ]
            )

            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.font = Font(name="Leelawadee")
                cell.border = border
                cell.alignment = center if c != 2 else left

                # 🎨 Daily coloring
                if daily_col_start <= c <= daily_col_end:
                    if v in ("", None):
                        cell.fill = blank_fill
                    elif str(v).upper() == "A":
                        cell.fill = absent_fill
                    elif str(v).upper() == "D":
                        cell.fill = dayoff_fill

            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        # ----------------------------
        # AUTO-FIT POSITION COLUMN WIDTH
        # ----------------------------
        # Adjust column 4 (Position) width based on max content length
        position_col_index = 4
        ws.column_dimensions[get_column_letter(position_col_index)].width = max_col_lengths[position_col_index - 1] + 2

       # ----------------------------
        # 🔢 GRAND TOTAL ROW (Per Column)
        # ----------------------------

        # Place "GRAND TOTAL" right before daily totals
        grand_total_col = daily_col_start - 1  # column just before daily totals
        top_left_col = grand_total_col - 1     # top-left of merged range (merge 2 columns for nicer display)

        # Merge two columns for GRAND TOTAL
        ws.merge_cells(
            start_row=row_idx, start_column=top_left_col,
            end_row=row_idx, end_column=grand_total_col
        )

        # Write text in top-left cell only
        cell = ws.cell(row=row_idx, column=top_left_col, value="GRAND TOTAL")
        cell.font = bold
        cell.alignment = Alignment(horizontal="right", vertical="center")

        # Apply border to all merged cells
        for col in range(top_left_col, grand_total_col + 1):
            ws.cell(row=row_idx, column=col).border = border

        right_start_col = daily_col_end + 1
        right_end_col = daily_col_end + len(headers_right)

        # Daily totals → count of numeric values
        for col in range(daily_col_start, daily_col_end + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = center
            if totals[col - 1] > 0:
                cell.value = totals[col - 1]  # count of numeric entries
                cell.font = bold

        # Right-side totals → sum numeric values
        for col in range(right_start_col, right_end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = center
            if totals[col - 1] > 0:
                cell.value = totals[col - 1]
                cell.font = bold

        ws.row_dimensions[row_idx].height = 22

        # ----------------------------
        # ✍ SIGNATURES ROWS
        # ----------------------------
        row_idx += 2  # leave one empty row below GRAND TOTAL

        # First row: Titles
        titles = ["Prepared By:", "Checked By:", "Noted By:"]
        col_positions = [2, 6, 12]  # B, F, L

        for col, title in zip(col_positions, titles):
            cell = ws.cell(row=row_idx, column=col, value=title)
            cell.font = bold
            cell.alignment = left

        sig_title_row = row_idx
        row_idx += 3

        # Second row: Signature lines
        for col in col_positions:
            cell = ws.cell(row=row_idx, column=col, value="___________________________")
            cell.font = bold
            cell.alignment = left

        row_idx += 1

        # Third row: Designations
        designations = ["Supervisor", "Operation Officer", "Client Representative"]
        for col, des in zip(col_positions, designations):
            cell = ws.cell(row=row_idx, column=col, value=des)
            cell.font = Font(name="Leelawadee", bold=True)
            cell.alignment = left

        # Adjust row heights
        ws.row_dimensions[sig_title_row].height = 20
        ws.row_dimensions[sig_title_row + 3].height = 20
        ws.row_dimensions[sig_title_row + 4].height = 20


        # ----------------------------
        # 📚 LEGENDS (Beside Noted By)
        # ----------------------------
        legend_start_row = sig_title_row
        legend_col = col_positions[2] + 13  # 3 columns to the right of "Noted By"

        # LEGENDS title
        cell = ws.cell(row=legend_start_row, column=legend_col, value="LEGENDS:")
        cell.font = bold
        cell.alignment = left

        legend_entries = [
            "A = ABSENT",
            "D = DAY-OFF",
            "L = LEAVE",
            "LR = LEAVE W/ RELIEVER",
            "RH = REGULAR HOLIDAY (8)",
            "RS = RESIGNED",
            "SH = SPECIAL HOLIDAY (8)",
            "T = TRANSFERRED",
            "R = RELIEVED",
            "S = SUSPENDED W/O RELIEVER",
            "SR = SUSPENDED W/ RELIEVER",
        ]

        for i, entry in enumerate(legend_entries, start=1):
            cell = ws.cell(row=legend_start_row + i, column=legend_col, value=entry)
            cell.font = Font(name="Leelawadee")
            cell.alignment = left


        # ----------------------------
        # 📏 AUTO WIDTH
        # ----------------------------
        for col_idx, max_len in enumerate(max_col_lengths, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(max_len + 2, 8)

        ws.freeze_panes = f"A{table_start + 2}"

        # ----------------------------
        # 📥 RESPONSE
        # ----------------------------
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="Attendance_{dtr.start_date}_to_{dtr.end_date}.xlsx"'
        )
        wb.save(response)
        return response

class DTREntryViewSet(viewsets.ModelViewSet):
    queryset = DTREntry.objects.all().order_by("full_name")
    serializer_class = DTREntrySerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"], url_path="employee")
    def by_employee(self, request):
        code = request.query_params.get("employee_code")

        if not code:
            return Response({"detail": "employee_code is required"}, status=400)

        normalized = code.lstrip("0").upper()

        qs = (
            self.get_queryset()
            .filter(
                Q(employee_no=code)
                | Q(employee_no=normalized)
                | Q(employee_no__endswith=normalized),
                dtr_file__status="verified",
            )
            .select_related("dtr_file__uploaded_by")  # ✅ important for performance
        )

        if not qs.exists():
            return Response([], status=200)

        grouped = {}
        for entry in qs:
            dtr_file = entry.dtr_file

            # ✅ Derive project name using same logic as sync_all_files
            uploader = getattr(dtr_file, "uploaded_by", None)
            project_name = (
                f"{uploader.first_name} {uploader.last_name}".strip()
                if uploader and (uploader.first_name or uploader.last_name)
                else uploader.username if uploader else "Unknown Project"
            )

            key = f"{dtr_file.start_date} → {dtr_file.end_date}"

            if key not in grouped:
                grouped[key] = {
                    "project": project_name,  # ✅ include here
                    "start_date": dtr_file.start_date,
                    "end_date": dtr_file.end_date,
                    "rows": [],
                }

            serialized_entry = DTREntrySerializer(entry).data
            serialized_entry["project"] = project_name  # ✅ ensure it’s attached per row too
            grouped[key]["rows"].append(serialized_entry)

        return Response(list(grouped.values()))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_basic_employees(request):
    employees = request.data.get("employees", [])
    if not employees:
        return Response({"detail": "No employees provided"}, status=status.HTTP_400_BAD_REQUEST)

    created, updated, skipped = 0, 0, 0

    for emp in employees:
        serializer = EmployeeSerializer(data=emp)
        if serializer.is_valid():
            emp_no = serializer.validated_data["employee_no"]
            emp_name = serializer.validated_data["employee_name"]

            obj, created_flag = Employee.objects.update_or_create(
                employee_no=emp_no,
                defaults={"employee_name": emp_name}
            )
            if created_flag:
                created += 1
            else:
                updated += 1
        else:
            skipped += 1

    return Response(
        {"detail": f"{created} created, {updated} updated, {skipped} skipped"},
        status=status.HTTP_200_OK
    )

logger = logging.getLogger(__name__)

def normalize_name(s):
    if not s:
        return ""
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip().upper()

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def compare_employees(request):
    """
    Compare EmployeeDirectory names to Employee (source of truth).
    Also removes duplicates (by employee_code or employee_name).
    Supports:
      - POST body { "threshold": 85, "dry_run": true/false }
      - GET query params ?threshold=85&dry_run=1 for quick tests
    """
    try:
        threshold = int(request.data.get("threshold") or request.query_params.get("threshold") or 85)
        dry_run_param = request.data.get("dry_run", request.query_params.get("dry_run", "false"))
        dry_run = str(dry_run_param).lower() in ("1", "true", "yes")

        employees_qs = Employee.objects.all()
        directories_qs = EmployeeDirectory.objects.all()

        duplicates_deleted = []
        seen_codes, seen_names = set(), set()
        for d in directories_qs:
            norm_name = normalize_name(d.employee_name)
            if d.employee_code in seen_codes or norm_name in seen_names:
                duplicates_deleted.append({
                    "employee_code": d.employee_code,
                    "employee_name": d.employee_name
                })
                if not dry_run:
                    d.delete()
            else:
                seen_codes.add(d.employee_code)
                seen_names.add(norm_name)

        directories_qs = EmployeeDirectory.objects.all()

        employees = []
        for emp in employees_qs:
            norm = normalize_name(emp.employee_name)
            employees.append({
                "obj": emp,
                "employee_no": emp.employee_no,
                "employee_name": emp.employee_name,
                "norm_name": norm
            })

        directories = []
        for d in directories_qs:
            directories.append({
                "obj": d,
                "employee_code": d.employee_code,
                "employee_name": d.employee_name,
                "norm_name": normalize_name(d.employee_name)
            })

        updated, added, skipped = [], [], []

        for d in directories:
            best_match = None
            best_score = 0
            for e in employees:
                score = fuzz.token_set_ratio(d["norm_name"], e["norm_name"])
                if score > best_score:
                    best_score = score
                    best_match = e

            if best_match and best_score >= threshold:
                obj = d["obj"]
                new_code = best_match["employee_no"]
                new_name = best_match["employee_name"]

                if not dry_run:
                    # Check for conflicts
                    conflict = EmployeeDirectory.objects.filter(employee_code=new_code).exclude(id=obj.id).exists()
                    if conflict:
                        skipped.append({
                            "code": d["employee_code"],
                            "name": d["employee_name"],
                            "reason": f"Target employee_code {new_code} already exists"
                        })
                    else:
                        obj.employee_code = new_code
                        obj.employee_name = new_name
                        obj.save()
                updated.append({
                    "old_code": d["employee_code"],
                    "old_name": d["employee_name"],
                    "new_code": new_code,
                    "new_name": new_name,
                    "match_score": best_score
                })

        existing_codes = set(EmployeeDirectory.objects.values_list("employee_code", flat=True))
        existing_norm_names = set(
            normalize_name(n) for n in EmployeeDirectory.objects.values_list("employee_name", flat=True)
        )

        for e in employees:
            if (e["employee_no"] not in existing_codes) and (e["norm_name"] not in existing_norm_names):
                added.append({
                    "new_code": e["employee_no"],
                    "new_name": e["employee_name"]
                })
                if not dry_run:
                    EmployeeDirectory.objects.create(
                        employee_code=e["employee_no"],
                        employee_name=e["employee_name"]
                    )
                    existing_codes.add(e["employee_no"])
                    existing_norm_names.add(e["norm_name"])

        summary = {
            "updated": len(updated),
            "added": len(added),
            "skipped": len(skipped),
            "duplicates_deleted": len(duplicates_deleted),
            "dry_run": dry_run,
            "threshold": threshold
        }

        return Response({
            "summary": summary,
            "updated_records": updated,
            "added_records": added,
            "skipped_records": skipped,
            "deleted_duplicates": duplicates_deleted
        }, status=status.HTTP_200_OK)

    except Exception as exc:
        logger.exception("compare_employees failed")
        return Response({
            "detail": "Internal error during compare",
            "error": str(exc)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(["POST"])
def flush_employee_data(request):
    """
    Safely reset all numeric/DTR fields in Employee model to None or 0.
    Only updates actual fields existing in the model.
    """
    flush_columns = request.data.get("flush_columns", [])

    valid_fields = []
    for col in flush_columns:
        try:
            EmployeeDirectory._meta.get_field(col)
            valid_fields.append(col)
        except FieldDoesNotExist:
            print(f"Skipping invalid field: {col}")  

    if not valid_fields:
        return Response({"detail": "No valid fields to flush."}, status=400)

    reset_data = {col: None for col in valid_fields}
    try:
        EmployeeDirectory.objects.all().update(**reset_data)
        return Response({"detail": f"Flushed fields: {valid_fields}"})
    except Exception as e:
        return Response({"detail": str(e)}, status=500)
    
@api_view(["GET"])
@permission_classes([IsAdminUser])
def backup_employee_directory(request):
    """
    Backup ALL EmployeeDirectory columns into an Excel file.
    """
    employees = EmployeeDirectory.objects.all().values()
    if not employees.exists():
        return Response({"detail": "No employee data found."}, status=404)

    df = pd.DataFrame(list(employees))

    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)
            except Exception:
                pass

    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)

    filename = f"employee_directory_backup_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@api_view(["POST"])
@permission_classes([IsAdminUser])
def restore_employee_directory(request):
    """
    Restore EmployeeDirectory data from an Excel backup.
    Includes ALL model fields.
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file uploaded."}, status=400)

    try:
        df = pd.read_excel(file).fillna("")

        if df.empty:
            return Response({"detail": "The backup file is empty."}, status=400)

        EmployeeDirectory.objects.all().delete()

        model_fields = [field.name for field in EmployeeDirectory._meta.get_fields() if field.concrete and not field.is_relation]
        objs = []

        for _, row in df.iterrows():
            field_data = {}
            for field in model_fields:
                if field in df.columns:
                    val = row[field]
                    if pd.isna(val) or val == "":
                        val = None
                    field_data[field] = val

            objs.append(EmployeeDirectory(**field_data))

        EmployeeDirectory.objects.bulk_create(objs)

        return Response({
            "detail": f"Restore complete. {len(objs)} records restored."
        }, status=200)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({
            "detail": f"Restore failed: {str(e)}"
        }, status=400)
    
class PDFFileViewSet(viewsets.ModelViewSet):
    serializer_class = PDFFileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return PDFFile.objects.all().order_by("-uploaded_at")
        return PDFFile.objects.filter(uploaded_by=user).order_by("-uploaded_at")

    @action(detail=True, methods=["put"], url_path="update-parsed")
    def update_parsed(self, request, pk=None):
        """Allow admins to edit and save parsed PDF data."""
        pdf = self.get_object()

        # ✅ Only superusers (admin role) can modify
        if not request.user.is_superuser:
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

        parsed_pages = request.data.get("parsed_pages")
        if not parsed_pages:
            return Response({"error": "Missing parsed_pages data"}, status=status.HTTP_400_BAD_REQUEST)

        pdf.parsed_pages = parsed_pages
        pdf.save()
        return Response({"message": "✅ Parsed data updated successfully!"})
